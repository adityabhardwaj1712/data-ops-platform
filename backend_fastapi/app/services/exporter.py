"""
Data Export Service

Handles exporting dataset versions to various formats (Excel, CSV, JSON)
with source links and confidence scores.
"""
import os
import uuid
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import ExportFormat, ExportRequest, ExportResponse


class DataExporter:
    """Service for exporting dataset versions to various formats"""

    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)

    async def export_dataset(
        self,
        data: List[Dict[str, Any]],
        request: ExportRequest,
        job_name: str = "dataset"
    ) -> ExportResponse:
        """
        Export dataset to the requested format

        Args:
            data: Dataset records to export
            request: Export configuration
            job_name: Name for the export file

        Returns:
            ExportResponse with file information
        """
        if not data:
            raise ValueError("No data to export")

        # Prepare data for export
        export_data = self._prepare_export_data(data, request)

        # Generate unique filename
        file_id = str(uuid.uuid4())[:8]
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        
        # Use consistent version number for both filename and response
        version_num = request.version or 1

        if request.format == ExportFormat.EXCEL:
            filename = f"{job_name}_v{version_num}_{timestamp}_{file_id}.xlsx"
            filepath = self.export_dir / filename
            await self._export_to_excel(export_data, filepath, request)
        elif request.format == ExportFormat.CSV:
            filename = f"{job_name}_v{version_num}_{timestamp}_{file_id}.csv"
            filepath = self.export_dir / filename
            await self._export_to_csv(export_data, filepath)
        elif request.format == ExportFormat.JSON:
            filename = f"{job_name}_v{version_num}_{timestamp}_{file_id}.json"
            filepath = self.export_dir / filename
            await self._export_to_json(export_data, filepath)
        else:
            raise ValueError(f"Unsupported export format: {request.format}")

        return ExportResponse(
            job_id=request.job_id,
            version=version_num,
            format=request.format,
            file_url=f"/exports/{filename}",
            row_count=len(export_data),
            created_at=datetime.utcnow()
        )

    def _prepare_export_data(
        self,
        data: List[Dict[str, Any]],
        request: ExportRequest
    ) -> List[Dict[str, Any]]:
        """Prepare data for export by flattening and adding metadata"""
        prepared_data = []

        for record in data:
            row = {}

            # Add main data fields
            for key, value in record.items():
                if key.startswith('_'):
                    continue  # Skip internal fields unless requested

                # Handle nested dictionaries/lists
                if isinstance(value, dict):
                    for nested_key, nested_value in value.items():
                        row[f"{key}_{nested_key}"] = nested_value
                elif isinstance(value, list):
                    row[key] = ', '.join(str(item) for item in value)
                else:
                    row[key] = value

            # Add source link if requested
            if request.include_source_links:
                source_url = record.get('_source_url') or record.get('source_url')
                if source_url:
                    row['Source_URL'] = source_url

            # Add confidence if requested
            if request.include_confidence:
                confidence = record.get('_confidence') or record.get('confidence')
                if confidence is not None:
                    row['Confidence_Score'] = confidence

                # Add field-level confidence if available
                field_confidence = record.get('field_confidence')
                if field_confidence:
                    for field, conf in field_confidence.items():
                        row[f'{field}_confidence'] = conf

            # Add extraction metadata
            extracted_at = record.get('_extracted_at') or record.get('extracted_at')
            if extracted_at:
                row['Extracted_At'] = extracted_at

            prepared_data.append(row)

        return prepared_data

    async def _export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filepath: Path,
        request: ExportRequest
    ):
        """Export data to Excel format with styling"""
        df = pd.DataFrame(data)

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write main data sheet
            df.to_excel(writer, sheet_name='Data', index=False)

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

                # Auto-adjust column width
                column_letter = get_column_letter(col_num)
                max_length = max(
                    len(str(column_title)),
                    max([len(str(df.iloc[i, col_num-1])) for i in range(min(len(df), 100))], default=0)
                )
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add metadata sheet if confidence data is included
            if request.include_confidence and any('confidence' in col.lower() for col in df.columns):
                confidence_cols = [col for col in df.columns if 'confidence' in col.lower()]
                if confidence_cols:
                    confidence_df = df[confidence_cols].copy()
                    confidence_df.insert(0, 'Row_Number', range(1, len(df) + 1))
                    confidence_df.to_excel(writer, sheet_name='Confidence', index=False)

                    # Style confidence sheet
                    conf_sheet = writer.sheets['Confidence']
                    for col_num, column_title in enumerate(confidence_df.columns, 1):
                        cell = conf_sheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill

    async def _export_to_csv(self, data: List[Dict[str, Any]], filepath: Path):
        """Export data to CSV format"""
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')

    async def _export_to_json(self, data: List[Dict[str, Any]], filepath: Path):
        """Export data to JSON format"""
        import json
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str, ensure_ascii=False)

    def cleanup_old_exports(self, max_age_days: int = 7):
        """Clean up old export files"""
        cutoff_time = datetime.utcnow().timestamp() - (max_age_days * 24 * 60 * 60)

        for file_path in self.export_dir.glob("*"):
            if file_path.is_file() and file_path.stat().st_mtime < cutoff_time:
                try:
                    file_path.unlink()
                except Exception:
                    pass  # Ignore cleanup errors


# Global instance
exporter = DataExporter()