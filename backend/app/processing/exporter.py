"""
Data Export Service

Handles exporting dataset versions to various formats (Excel, CSV, JSON)
with source links and confidence scores.
"""
import os
import uuid
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import json
import zipfile
import shutil

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import ExportFormat, ExportRequest, ExportResponse


class DataExporter:
    """Service for exporting dataset versions to various formats"""

    def __init__(self, export_dir: str = "/app/data/exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    async def export_dataset(
        self,
        data: List[Dict[str, Any]],
        request: ExportRequest,
        job_name: str = "dataset"
    ) -> ExportResponse:
        """
        Export dataset to the requested format

        Args:
            data: Dataset records to export
            request: Export configuration
            job_name: Name for the export file

        Returns:
            ExportResponse with file information
        """
        if not data:
            raise ValueError("No data to export")

        # Prepare data for export
        export_data = self._prepare_export_data(data, request)

        # Generate unique filename
        file_id = str(uuid.uuid4())[:8]
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        
        # Use consistent version number for both filename and response
        version_num = request.version or 1

        if request.format == ExportFormat.EXCEL:
            filename = f"{job_name}_v{version_num}_{timestamp}_{file_id}.xlsx"
            filepath = self.export_dir / filename
            await self._export_to_excel(export_data, filepath, request)
        elif request.format == ExportFormat.CSV:
            filename = f"{job_name}_v{version_num}_{timestamp}_{file_id}.csv"
            filepath = self.export_dir / filename
            await self._export_to_csv(export_data, filepath)
        elif request.format == ExportFormat.JSON:
            filename = f"{job_name}_v{version_num}_{timestamp}_{file_id}.json"
            filepath = self.export_dir / filename
            await self._export_to_json(export_data, filepath)
        else:
            raise ValueError(f"Unsupported export format: {request.format}")

        return ExportResponse(
            job_id=request.job_id,
            version=version_num,
            format=request.format,
            file_url=f"/exports/{filename}",
            row_count=len(export_data),
            created_at=datetime.utcnow()
        )

    async def create_client_package(
        self,
        data: List[Dict[str, Any]],
        request: ExportRequest,
        job_name: str,
        metadata: Dict[str, Any],
        artifact_paths: Optional[List[str]] = None
    ) -> ExportResponse:
        """
        Create a professional ZIP package for client delivery.
        Includes: data file, metadata.json, and README.txt
        """
        version_num = request.version or 1
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        package_name = f"{job_name}_delivery_v{version_num}_{timestamp}"
        temp_dir = self.export_dir / package_name
        temp_dir.mkdir(exist_ok=True)

        # 1. Export Data File
        data_filename = f"{job_name}_data.{'xlsx' if request.format == ExportFormat.EXCEL else 'csv'}"
        data_path = temp_dir / data_filename
        export_data = self._prepare_export_data(data, request)
        
        if request.format == ExportFormat.EXCEL:
            await self._export_to_excel(export_data, data_path, request)
        else:
            await self._export_to_csv(export_data, data_path)

        # 2. Export Metadata JSON
        meta_filename = "metadata.json"
        meta_path = temp_dir / meta_filename
        clean_meta = {
            "job_id": str(request.job_id),
            "job_name": job_name,
            "version": version_num,
            "row_count": len(export_data),
            "scrape_date": datetime.utcnow().isoformat(),
            "confidence_metrics": metadata.get("confidence_summary", {}),
            "human_reviewed": metadata.get("is_human", False)
        }
        with open(meta_path, "w") as f:
            json.dump(clean_meta, f, indent=2)

        # 3. Handle Artifacts (Screenshots, HTML)
        if artifact_paths:
            artifacts_dir = temp_dir / "artifacts"
            artifacts_dir.mkdir(exist_ok=True)
            for ap in artifact_paths:
                ap_path = Path(ap)
                if ap_path.exists():
                    shutil.copy2(ap_path, artifacts_dir / ap_path.name)

        # 4. Create README.txt
        readme_path = temp_dir / "README.txt"
        self._generate_readme(readme_path, job_name, export_data, clean_meta, bool(artifact_paths))

        # 4. Zip it all up
        zip_filename = f"{package_name}.zip"
        zip_path = self.export_dir / zip_filename
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in temp_dir.glob("*"):
                zipf.write(file, arcname=file.name)

        # Cleanup temp dir
        shutil.rmtree(temp_dir)

        return ExportResponse(
            job_id=request.job_id,
            version=version_num,
            format=request.format,
            file_url=f"/exports/{zip_filename}",
            row_count=len(export_data),
            created_at=datetime.utcnow()
        )

    def _generate_readme(self, path: Path, job_name: str, data: List[Dict[str, Any]], meta: Dict[str, Any], has_artifacts: bool = False):
        """Generate a professional README from template."""
        template_path = Path(__file__).parent.parent / "templates" / "README_DELIVERY.txt"
        
        if not template_path.exists():
            content = f"Data Delivery for {job_name}\nDate: {meta['scrape_date']}\nRows: {meta['row_count']}"
        else:
            with open(template_path, "r") as f:
                content = f.read()

        # Simple field definitions
        fields = data[0].keys() if data else []
        field_defs = "\n".join([f"- {f}: Extracted data field" for f in fields if not f.startswith('_')])

        content = content.format(
            job_name=job_name,
            delivery_date=meta['scrape_date'][:10],
            source_url="Provided in CSV", # Or pull from meta if available
            row_count=meta['row_count'],
            confidence_score=meta['confidence_metrics'].get('score', 0),
            field_definitions=field_defs,
            hitl_status="Human reviewed for high accuracy." if meta['human_reviewed'] else "Automated high-confidence extraction.",
            artifact_note="Refer to the 'artifacts/' folder for raw HTML and screenshots." if has_artifacts else "Raw artifacts were not included in this package."
        )

        with open(path, "w") as f:
            f.write(content)

    def _prepare_export_data(
        self,
        data: List[Dict[str, Any]],
        request: ExportRequest
    ) -> List[Dict[str, Any]]:
        """Prepare data for export by flattening and adding metadata"""
        prepared_data = []

        for record in data:
            row = {}

            # Add main data fields
            for key, value in record.items():
                if key.startswith('_'):
                    continue  # Skip internal fields unless requested

                # Handle nested dictionaries/lists
                if isinstance(value, dict):
                    for nested_key, nested_value in value.items():
                        row[f"{key}_{nested_key}"] = nested_value
                elif isinstance(value, list):
                    row[key] = ', '.join(str(item) for item in value)
                else:
                    row[key] = value

            # --- MANDATORY AUDIT FIELDS (Sprint 1 Requirement) ---
            # 1. Source URL
            row['source_url'] = record.get('_source_url') or record.get('source_url') or record.get('metadata', {}).get('url', 'N/A')
            
            # 2. Scrape Timestamp
            row['scrape_timestamp'] = record.get('_extracted_at') or record.get('extracted_at') or record.get('metadata', {}).get('timestamp', datetime.utcnow().isoformat())
            
            # 3. Confidence Score
            row['confidence_score'] = record.get('_confidence') or record.get('confidence') or record.get('metadata', {}).get('confidence', 0.0)
            
            # 4. Human Reviewed
            row['reviewed'] = record.get('human_reviewed') or record.get('metadata', {}).get('human_reviewed', False)

            # Add field-level confidence if available
            field_confidence = record.get('field_confidence')
            if field_confidence:
                for field, conf in field_confidence.items():
                    row[f'{field}_confidence'] = conf

            prepared_data.append(row)

        return prepared_data

    async def _export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filepath: Path,
        request: ExportRequest
    ):
        """Export data to Excel format with styling"""
        df = pd.DataFrame(data)

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write main data sheet
            df.to_excel(writer, sheet_name='Data', index=False)

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

                # Auto-adjust column width
                column_letter = get_column_letter(col_num)
                max_length = max(
                    len(str(column_title)),
                    max([len(str(df.iloc[i, col_num-1])) for i in range(min(len(df), 100))], default=0)
                )
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add metadata sheet if confidence data is included
            if request.include_confidence and any('confidence' in col.lower() for col in df.columns):
                confidence_cols = [col for col in df.columns if 'confidence' in col.lower()]
                if confidence_cols:
                    confidence_df = df[confidence_cols].copy()
                    confidence_df.insert(0, 'Row_Number', range(1, len(df) + 1))
                    confidence_df.to_excel(writer, sheet_name='Confidence', index=False)

                    # Style confidence sheet
                    conf_sheet = writer.sheets['Confidence']
                    for col_num, column_title in enumerate(confidence_df.columns, 1):
                        cell = conf_sheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill

    async def _export_to_csv(self, data: List[Dict[str, Any]], filepath: Path):
        """Export data to CSV format"""
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')

    async def _export_to_json(self, data: List[Dict[str, Any]], filepath: Path):
        """Export data to JSON format"""
        import json
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str, ensure_ascii=False)

    def cleanup_old_exports(self, max_age_days: int = 7):
        """Clean up old export files"""
        cutoff_time = datetime.utcnow().timestamp() - (max_age_days * 24 * 60 * 60)

        for file_path in self.export_dir.glob("*"):
            if file_path.is_file() and file_path.stat().st_mtime < cutoff_time:
                try:
                    file_path.unlink()
                except Exception:
                    pass  # Ignore cleanup errors


# Global instance
exporter = DataExporter()